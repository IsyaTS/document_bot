from __future__ import annotations

import csv
import io
from decimal import Decimal
from typing import Any

from aiogram import Bot, Dispatcher, F, Router
from aiogram.exceptions import TelegramBadRequest
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import CallbackQuery, Document as TgDocument, FSInputFile, InlineKeyboardButton, InlineKeyboardMarkup, Message

from app.catalog import catalog_text, enrich_items_from_catalog, search_text, update_catalog_price
from app.company import Company, load_companies, requisites_text
from app.config import Settings
from app.documents import DOC_TYPES, generate_document
from app.moysklad import MoySkladClient
from app.openai_assistant import OpenAIDrafter
from app.parser import apply_discount, parse_items, parse_user_input
from app.pdf import generate_pdf_document
from app.storage import Storage


DOC_ICONS = {
    "door_offer": "🚪",
    "invoice": "💳",
    "measurement_estimate": "📐",
    "door_contract": "📝",
    "contract": "📝",
    "measurement_act": "📏",
    "installation_act": "🔧",
    "warranty": "🛡️",
    "offer": "📄",
    "claim_reply": "⚖️",
    "act": "✅",
    "waybill": "🚚",
    "reconciliation": "📊",
    "official_letter": "✉️",
}

DOOR_DOC_TYPES = {"door_offer", "measurement_estimate", "door_contract", "measurement_act", "installation_act", "warranty"}
DOOR_ORDER_DOC_TYPES = {"door_offer", "measurement_estimate", "door_contract", "warranty"}
DOOR_ACCEPTANCE_DOC_TYPES = {"measurement_act", "installation_act"}

TEMPLATE_PRESETS = {
    "economy": {"label": "Эконом", "template": "эконом", "model": "Входная дверь эконом"},
    "standard": {"label": "Стандарт", "template": "стандарт", "model": "Входная дверь стандарт"},
    "premium": {"label": "Премиум", "template": "премиум", "model": "Входная дверь премиум"},
    "interior": {"label": "Межкомнатная", "template": "межкомнатная", "model": "Межкомнатная дверь"},
}

FIELD_LABELS = {
    "counterparty_name": "Клиент",
    "counterparty_inn": "ИНН",
    "counterparty_kpp": "КПП",
    "counterparty_ogrn": "ОГРН",
    "counterparty_address": "Адрес",
    "counterparty_phone": "Телефон",
    "counterparty_email": "Email",
    "counterparty_manager": "Представитель",
    "subject": "Предмет",
    "basis": "Основание",
    "template": "Шаблон",
    "model": "Модель",
    "object_address": "Адрес объекта",
    "opening_size": "Размер проема",
    "wall_depth": "Толщина стены",
    "color": "Цвет",
    "opening_side": "Открывание",
    "term": "Срок",
    "payment_terms": "Оплата",
    "payment_purpose": "Назначение платежа",
    "vat": "НДС",
    "discount": "Скидка",
    "delivery": "Доставка",
    "installation": "Монтаж",
    "dismantling": "Демонтаж",
    "lifting": "Подъем",
    "warranty_period": "Гарантия",
    "measurer": "Замерщик",
    "serial_number": "Серийный номер",
    "route": "Маршрут",
    "loading_address": "Адрес погрузки",
    "unloading_address": "Адрес выгрузки",
    "claim_text": "Претензия",
    "description": "Комментарий",
    "number": "Номер",
}

FIELD_ORDER = [
    "counterparty_name",
    "counterparty_inn",
    "counterparty_kpp",
    "counterparty_ogrn",
    "counterparty_phone",
    "counterparty_email",
    "counterparty_manager",
    "counterparty_address",
    "object_address",
    "template",
    "model",
    "opening_size",
    "wall_depth",
    "color",
    "opening_side",
    "subject",
    "basis",
    "route",
    "loading_address",
    "unloading_address",
    "term",
    "payment_terms",
    "payment_purpose",
    "vat",
    "discount",
    "warranty_period",
    "measurer",
    "serial_number",
    "delivery",
    "installation",
    "dismantling",
    "lifting",
    "claim_text",
    "description",
    "number",
]


class DocumentFlow(StatesGroup):
    choosing_doc_type = State()
    choosing_company = State()
    choosing_mode = State()
    choosing_saved_counterparty = State()
    waiting_details = State()
    waiting_repeat_changes = State()
    waiting_product_search = State()
    waiting_price_update = State()
    waiting_quick_input = State()
    waiting_edit_value = State()
    waiting_counterparty_manual = State()
    waiting_counterparty_import = State()


def build_dispatcher(settings: Settings) -> Dispatcher:
    companies = load_companies()
    storage = Storage(settings.database_path)
    drafter = OpenAIDrafter(settings.openai_api_key, settings.openai_model)
    moysklad = MoySkladClient(settings.moysklad_login, settings.moysklad_password)

    router = Router()

    @router.message(Command("start", "menu", "doc", "new", "group"))
    async def start(message: Message, state: FSMContext) -> None:
        await state.clear()
        await message.answer(
            "Бот для дверей и документооборота.\n"
            "Есть быстрый мастер, полный ввод, подстановка из прошлого документа и сохраненные контрагенты.",
            reply_markup=_main_keyboard(),
        )

    @router.message(Command("id", "chatid"))
    async def chat_id(message: Message) -> None:
        await message.answer(_chat_id_text(message))

    @router.message(Command("help"))
    async def help_message(message: Message) -> None:
        await message.answer(_help_text())

    @router.message(Command("requisites"))
    async def requisites(message: Message) -> None:
        await message.answer("Выберите организацию:", reply_markup=_companies_keyboard(companies, prefix="req"))

    @router.message(Command("history"))
    async def history(message: Message) -> None:
        rows = storage.recent_documents(message.from_user.id)
        if not rows:
            await message.answer("История пустая.")
            return
        text = "\n".join(f"#{doc_id} {created_at}: {DOC_TYPES.get(doc_type, doc_type)} - {filename}" for doc_id, doc_type, filename, created_at in rows)
        await message.answer(text)

    @router.message(Command("counterparties"))
    async def counterparties_command(message: Message) -> None:
        rows = storage.recent_counterparties(message.from_user.id)
        text = "Сохраненные контрагенты:" if rows else "Сохраненных контрагентов пока нет."
        await message.answer(text, reply_markup=_counterparties_menu_keyboard(rows))

    @router.message(Command("catalog"))
    async def catalog_command(message: Message) -> None:
        query = _command_args(message)
        if query:
            await message.answer(_product_search_text(query, moysklad), reply_markup=_main_keyboard())
            return
        await message.answer(catalog_text(), reply_markup=_catalog_keyboard())

    @router.message(Command("product"))
    async def product_command(message: Message) -> None:
        query = _command_args(message)
        if not query:
            await message.answer("Напишите: /product часть названия товара")
            return
        await message.answer(_product_search_text(query, moysklad), reply_markup=_catalog_keyboard())

    @router.message(Command("moysklad"))
    async def moysklad_search(message: Message) -> None:
        query = _command_args(message)
        if not query:
            await message.answer("Напишите: /moysklad название или ИНН контрагента")
            return
        if not moysklad.enabled:
            await message.answer("Интеграция с МойСклад не настроена. Добавьте MOYSKLAD_LOGIN и MOYSKLAD_PASSWORD в .env.")
            return
        try:
            rows = moysklad.search_counterparty(query)
        except Exception as exc:
            await message.answer(f"МойСклад вернул ошибку: {exc}")
            return
        if not rows:
            await message.answer("Контрагент не найден.")
            return
        lines = []
        for row in rows:
            lines.append(f"{row.get('name', '-')}; ИНН: {row.get('inn', '-')}; КПП: {row.get('kpp', '-')}")
        await message.answer("\n".join(lines))

    @router.callback_query(F.data == "create")
    async def create(callback: CallbackQuery, state: FSMContext) -> None:
        await state.set_state(DocumentFlow.choosing_doc_type)
        await _show_callback_screen(callback, "Выберите тип документа:", reply_markup=_doc_types_keyboard())
        await _safe_callback_answer(callback)

    @router.callback_query(F.data == "catalog")
    async def catalog(callback: CallbackQuery) -> None:
        await _show_callback_screen(callback, catalog_text(), reply_markup=_catalog_keyboard())
        await _safe_callback_answer(callback)

    @router.callback_query(F.data == "counterparties")
    async def counterparties_menu(callback: CallbackQuery) -> None:
        rows = storage.recent_counterparties(callback.from_user.id)
        text = "Сохраненные контрагенты:" if rows else "Сохраненных контрагентов пока нет."
        await _show_callback_screen(callback, text, reply_markup=_counterparties_menu_keyboard(rows))
        await _safe_callback_answer(callback)

    @router.callback_query(F.data == "cpadd")
    async def add_counterparty_prompt(callback: CallbackQuery, state: FSMContext) -> None:
        await state.set_state(DocumentFlow.waiting_counterparty_manual)
        await _show_callback_screen(
            callback,
            "Пришлите контрагента одним сообщением.\n\n"
            "Пример:\n"
            "Клиент: ООО Ромашка\n"
            "ИНН: 1234567890\n"
            "КПП: 123401001\n"
            "Адрес: г. Уфа, ул. Примерная, 1\n"
            "Телефон: +7 900 000-00-00\n"
            "Email: info@romashka.ru\n"
            "Представитель: Иванов И.И.",
            reply_markup=_cancel_keyboard(),
        )
        await _safe_callback_answer(callback)

    @router.callback_query(F.data == "cpimport")
    async def import_counterparty_prompt(callback: CallbackQuery, state: FSMContext) -> None:
        await state.set_state(DocumentFlow.waiting_counterparty_import)
        await _show_callback_screen(
            callback,
            "Пришлите список контрагентов.\n\n"
            "Варианты:\n"
            "1. Вставить текстом несколькими блоками, разделяя пустой строкой.\n"
            "2. Отправить файл `.csv`, `.txt` или `.xlsx`.\n\n"
            "Колонки для CSV/XLSX:\n"
            "name, inn, kpp, ogrn, address, phone, email, manager",
            reply_markup=_cancel_keyboard(),
        )
        await _safe_callback_answer(callback)

    @router.callback_query(F.data.startswith("cpview:"))
    async def counterparty_view(callback: CallbackQuery) -> None:
        counterparty_id = int(callback.data.split(":", 1)[1])
        row = storage.get_counterparty(callback.from_user.id, counterparty_id)
        if not row:
            await _safe_callback_answer(callback, "Контрагент не найден", show_alert=True)
            return
        await _show_callback_screen(callback, _counterparty_text(row), reply_markup=_main_keyboard())
        await _safe_callback_answer(callback)

    @router.message(DocumentFlow.waiting_counterparty_manual, F.text)
    async def receive_counterparty_manual(message: Message, state: FSMContext) -> None:
        parsed = parse_user_input(message.text or "")
        counterparty_id = storage.upsert_counterparty(message.from_user.id, parsed.fields)
        if not counterparty_id:
            await message.answer("Не смог сохранить контрагента. Нужен хотя бы `Клиент:`.", reply_markup=_cancel_keyboard())
            return
        await state.clear()
        rows = storage.recent_counterparties(message.from_user.id)
        await message.answer("Контрагент сохранен.", reply_markup=_counterparties_menu_keyboard(rows))

    @router.message(DocumentFlow.waiting_counterparty_import, F.text)
    async def receive_counterparty_import_text(message: Message, state: FSMContext) -> None:
        count = _import_counterparties_from_text(storage, message.from_user.id, message.text or "")
        if count <= 0:
            await message.answer("Ничего не импортировал. Проверь формат блоков или CSV-строк.", reply_markup=_cancel_keyboard())
            return
        await state.clear()
        rows = storage.recent_counterparties(message.from_user.id)
        await message.answer(f"Импорт завершен. Сохранено: {count}.", reply_markup=_counterparties_menu_keyboard(rows))

    @router.message(DocumentFlow.waiting_counterparty_import, F.document)
    async def receive_counterparty_import_file(message: Message, state: FSMContext) -> None:
        document = message.document
        if document is None:
            return
        name = (document.file_name or "").lower()
        if name.endswith(".csv") or name.endswith(".txt"):
            text = await _download_text_file(message.bot, document)
            if text is None:
                await message.answer("Не смог прочитать файл.", reply_markup=_cancel_keyboard())
                return
            count = _import_counterparties_from_text(storage, message.from_user.id, text, prefer_csv=name.endswith(".csv"))
        elif name.endswith(".xlsx"):
            binary = await _download_binary_file(message.bot, document)
            if binary is None:
                await message.answer("Не смог прочитать Excel-файл.", reply_markup=_cancel_keyboard())
                return
            count = _import_counterparties_from_xlsx(storage, message.from_user.id, binary)
        else:
            await message.answer("Поддерживаются `.csv`, `.txt` и `.xlsx`.", reply_markup=_cancel_keyboard())
            return
        if count <= 0:
            await message.answer("Файл прочитан, но контрагенты не найдены. Проверьте формат.", reply_markup=_cancel_keyboard())
            return
        await state.clear()
        rows = storage.recent_counterparties(message.from_user.id)
        await message.answer(f"Импорт завершен. Сохранено: {count}.", reply_markup=_counterparties_menu_keyboard(rows))

    @router.callback_query(F.data == "search_product")
    async def search_product(callback: CallbackQuery, state: FSMContext) -> None:
        await state.set_state(DocumentFlow.waiting_product_search)
        await _show_callback_screen(
            callback,
            "Напишите часть названия товара или услуги: стандарт, прем, монтаж, доставка.",
            reply_markup=_cancel_keyboard(),
        )
        await _safe_callback_answer(callback)

    @router.message(DocumentFlow.waiting_product_search)
    async def receive_product_search(message: Message, state: FSMContext) -> None:
        await message.answer(_product_search_text(message.text or "", moysklad), reply_markup=_catalog_keyboard())
        await state.clear()

    @router.callback_query(F.data == "edit_price")
    async def edit_price(callback: CallbackQuery, state: FSMContext) -> None:
        await state.set_state(DocumentFlow.waiting_price_update)
        await _show_callback_screen(
            callback,
            "Напишите название и новую цену.\n\n"
            "Пример:\n"
            "стандарт: 44500\n"
            "или\n"
            "монтаж входной: 7000",
            reply_markup=_cancel_keyboard(),
        )
        await _safe_callback_answer(callback)

    @router.message(DocumentFlow.waiting_price_update)
    async def receive_price_update(message: Message, state: FSMContext) -> None:
        text = message.text or ""
        if ":" not in text:
            await message.answer("Нужен формат: название: цена", reply_markup=_catalog_keyboard())
            await state.clear()
            return
        name, raw_price = text.split(":", 1)
        price = _parse_price(raw_price)
        if price <= 0:
            await message.answer("Цена должна быть больше нуля.", reply_markup=_catalog_keyboard())
            await state.clear()
            return
        updated = update_catalog_price(name, price)
        if updated:
            await message.answer(f"Цена обновлена: {updated} - {price} руб.", reply_markup=_catalog_keyboard())
        else:
            await message.answer("Не нашел позицию. Попробуйте через поиск товара.", reply_markup=_catalog_keyboard())
        await state.clear()

    @router.callback_query(F.data == "history")
    async def history_button(callback: CallbackQuery) -> None:
        rows = storage.recent_documents(callback.from_user.id)
        if not rows:
            await _show_callback_screen(callback, "История пустая.", reply_markup=_main_keyboard())
            await _safe_callback_answer(callback)
            return
        text = "\n".join(f"#{doc_id} {created_at}: {DOC_TYPES.get(doc_type, doc_type)} - {filename}" for doc_id, doc_type, filename, created_at in rows)
        await _show_callback_screen(callback, text, reply_markup=_main_keyboard())
        await _safe_callback_answer(callback)

    @router.callback_query(F.data == "download_docs")
    async def download_docs(callback: CallbackQuery) -> None:
        rows = storage.recent_documents(callback.from_user.id, limit=10)
        if not rows:
            await _show_callback_screen(callback, "История пустая, скачивать пока нечего.", reply_markup=_main_keyboard())
            await _safe_callback_answer(callback)
            return
        await _show_callback_screen(callback, "Выберите документ для скачивания:", reply_markup=_download_keyboard(rows))
        await _safe_callback_answer(callback)

    @router.callback_query(F.data.startswith("dl:"))
    async def download_document(callback: CallbackQuery) -> None:
        document_id = int(callback.data.split(":", 1)[1])
        document = storage.get_document(callback.from_user.id, document_id)
        if not document:
            await callback.message.answer("Документ не найден.", reply_markup=_main_keyboard())
            await _safe_callback_answer(callback)
            return
        path = settings.documents_dir / document["filename"]
        if not path.exists():
            await callback.message.answer("Файл не найден на диске.", reply_markup=_main_keyboard())
            await _safe_callback_answer(callback)
            return
        await callback.message.answer_document(FSInputFile(path), caption=f"Документ #{document_id}", reply_markup=_main_keyboard())
        await _safe_callback_answer(callback)

    @router.callback_query(F.data == "repeat_offer")
    async def repeat_offer(callback: CallbackQuery, state: FSMContext) -> None:
        rows = storage.recent_documents(callback.from_user.id, limit=5, doc_type="door_offer")
        if not rows:
            rows = storage.recent_documents(callback.from_user.id, limit=5, doc_type="offer")
        if not rows:
            await _show_callback_screen(callback, "Пока нет КП для повтора. Создайте первое КП по дверям.", reply_markup=_main_keyboard())
            await _safe_callback_answer(callback)
            return
        await state.set_state(DocumentFlow.waiting_repeat_changes)
        await _show_callback_screen(callback, "Выберите КП, которое нужно повторить:", reply_markup=_repeat_keyboard(rows))
        await _safe_callback_answer(callback)

    @router.callback_query(F.data.startswith("rep:"))
    async def choose_repeat(callback: CallbackQuery, state: FSMContext) -> None:
        document_id = int(callback.data.split(":", 1)[1])
        document = storage.get_document(callback.from_user.id, document_id)
        if not document or not document["raw_text"]:
            await _show_callback_screen(callback, "Не нашел исходные данные этого КП. Создайте новое КП.", reply_markup=_main_keyboard())
            await _safe_callback_answer(callback)
            return
        await state.update_data(
            repeat_document_id=document_id,
            base_text=document["raw_text"],
            doc_type="door_offer",
            company_key=document["company_key"],
        )
        await state.set_state(DocumentFlow.waiting_repeat_changes)
        await _show_callback_screen(
            callback,
            "Пришлите только изменения. Можно написать `ок`, если нужен такой же документ.\n\n"
            "Пример:\n"
            "Скидка: 7%\n"
            "Позиции: Монтаж входной двери | 1 | шт. | 6500\n"
            "Срок: монтаж на этой неделе",
            reply_markup=_cancel_keyboard(),
        )
        await _safe_callback_answer(callback)

    @router.callback_query(F.data.startswith("doc:"))
    async def choose_doc(callback: CallbackQuery, state: FSMContext) -> None:
        doc_type = callback.data.split(":", 1)[1]
        await state.set_state(DocumentFlow.choosing_company)
        await state.update_data(
            doc_type=doc_type,
            draft_fields={},
            draft_items=[],
            selected_counterparty_id=None,
            quick_step_index=0,
        )
        await _show_callback_screen(
            callback,
            f"Выбрано: {DOC_TYPES.get(doc_type, doc_type)}.\nОт какой организации делаем документ?",
            reply_markup=_companies_keyboard(companies, prefix="company"),
        )
        await _safe_callback_answer(callback)

    @router.callback_query(F.data.startswith("company:"))
    async def choose_company(callback: CallbackQuery, state: FSMContext) -> None:
        company_key = callback.data.split(":", 1)[1]
        await state.update_data(company_key=company_key)
        await state.set_state(DocumentFlow.choosing_mode)
        data = await state.get_data()
        await _show_callback_screen(
            callback,
            _mode_prompt(data.get("doc_type", ""), bool(storage.last_document_snapshot(callback.from_user.id, data.get("doc_type", ""))), bool(storage.recent_counterparties(callback.from_user.id, limit=1))),
            reply_markup=_mode_keyboard(data.get("doc_type", ""), has_last=bool(storage.last_document_snapshot(callback.from_user.id, data.get("doc_type", ""))), has_counterparties=bool(storage.recent_counterparties(callback.from_user.id, limit=1))),
        )
        await _safe_callback_answer(callback)

    @router.callback_query(F.data.startswith("mode:"))
    async def choose_mode(callback: CallbackQuery, state: FSMContext) -> None:
        mode = callback.data.split(":", 1)[1]
        data = await state.get_data()
        doc_type = data.get("doc_type", "")

        if mode == "quick":
            await state.set_state(DocumentFlow.waiting_quick_input)
            await state.update_data(quick_step_index=0)
            await _prompt_quick_step(callback=callback, state=state)
            await _safe_callback_answer(callback)
            return

        if mode == "full":
            await state.set_state(DocumentFlow.waiting_details)
            seeded = _snapshot_to_raw_text({"fields": data.get("draft_fields", {}), "items": data.get("draft_items", [])})
            text = _details_prompt(doc_type, bool(data.get("draft_fields")))
            if seeded:
                text += "\n\nКонтрагент уже подставлен. Можно прислать только отличающиеся поля: адрес, позиции, оплату, срок."
            await _show_callback_screen(callback, text, reply_markup=_cancel_keyboard())
            await _safe_callback_answer(callback)
            return

        if mode == "reuse":
            last_doc = storage.last_document_snapshot(callback.from_user.id, doc_type) or storage.last_document_snapshot(callback.from_user.id)
            if not last_doc or not last_doc["raw_text"]:
                await _safe_callback_answer(callback, "Прошлого документа для подстановки пока нет.", show_alert=True)
                return
            await state.update_data(base_text=last_doc["raw_text"], company_key=data.get("company_key") or last_doc["company_key"])
            await state.set_state(DocumentFlow.waiting_repeat_changes)
            await _show_callback_screen(
                callback,
                "Подтянул данные из последнего документа. Пришлите только изменения или напишите `ок`, если нужно сформировать без изменений.",
                reply_markup=_cancel_keyboard(),
            )
            await _safe_callback_answer(callback)
            return

        rows = storage.recent_counterparties(callback.from_user.id)
        if not rows:
            await _safe_callback_answer(callback, "Сохраненных контрагентов пока нет.", show_alert=True)
            return
        await state.set_state(DocumentFlow.choosing_saved_counterparty)
        await _show_callback_screen(callback, "Выберите сохраненного контрагента:", reply_markup=_counterparties_keyboard(rows, prefix="cpuse"))
        await _safe_callback_answer(callback)

    @router.callback_query(F.data.startswith("cpuse:"))
    async def use_saved_counterparty(callback: CallbackQuery, state: FSMContext) -> None:
        counterparty_id = int(callback.data.split(":", 1)[1])
        row = storage.get_counterparty(callback.from_user.id, counterparty_id)
        if not row:
            await _safe_callback_answer(callback, "Контрагент не найден", show_alert=True)
            return
        draft_fields = _counterparty_to_fields(row)
        await state.update_data(draft_fields=draft_fields, selected_counterparty_id=counterparty_id)
        await state.set_state(DocumentFlow.choosing_mode)
        data = await state.get_data()
        await _show_callback_screen(
            callback,
            f"Контрагент подставлен: {row['name']}.\nТеперь выберите режим ввода.",
            reply_markup=_mode_keyboard(data.get("doc_type", ""), has_last=bool(storage.last_document_snapshot(callback.from_user.id, data.get("doc_type", ""))), has_counterparties=True),
        )
        await _safe_callback_answer(callback)

    @router.callback_query(F.data.startswith("req:"))
    async def company_requisites(callback: CallbackQuery) -> None:
        company_key = callback.data.split(":", 1)[1]
        company = companies[company_key]
        await _show_callback_screen(callback, requisites_text(company), reply_markup=_main_keyboard())
        await _safe_callback_answer(callback)

    @router.callback_query(F.data.startswith("quickv:"))
    async def quick_callback_value(callback: CallbackQuery, state: FSMContext) -> None:
        data = await state.get_data()
        if data.get("quick_step_index") is None:
            await _safe_callback_answer(callback)
            return
        step = _current_quick_step(data)
        if not step:
            await _safe_callback_answer(callback)
            return
        value = callback.data.split(":", 1)[1]
        await _consume_quick_value(
            callback=callback,
            state=state,
            step=step,
            raw_value=value,
            storage=storage,
            drafter=drafter,
            settings=settings,
            companies=companies,
        )
        await _safe_callback_answer(callback)

    @router.message(DocumentFlow.waiting_quick_input)
    async def receive_quick_input(message: Message, state: FSMContext) -> None:
        data = await state.get_data()
        step = _current_quick_step(data)
        if not step:
            await state.clear()
            await message.answer("Быстрый мастер сбросился. Нажмите /menu.", reply_markup=_main_keyboard())
            return
        await _consume_quick_value(
            message=message,
            state=state,
            step=step,
            raw_value=message.text or "",
            storage=storage,
            drafter=drafter,
            settings=settings,
            companies=companies,
        )

    @router.message(DocumentFlow.waiting_details)
    async def receive_details(message: Message, state: FSMContext) -> None:
        data = await state.get_data()
        doc_type = data["doc_type"]
        company = companies[data["company_key"]]
        seed_raw = _snapshot_to_raw_text({"fields": data.get("draft_fields", {}), "items": data.get("draft_items", [])})
        raw_text = _merge_raw_text(seed_raw, message.text or "")
        parsed = _prepare_parsed(raw_text)
        await _generate_and_send(
            message,
            storage,
            drafter,
            settings,
            doc_type,
            company,
            parsed,
            raw_text,
            selected_counterparty_id=data.get("selected_counterparty_id"),
        )
        await state.clear()

    @router.message(DocumentFlow.waiting_repeat_changes)
    async def receive_repeat_changes(message: Message, state: FSMContext) -> None:
        data = await state.get_data()
        base_text = data.get("base_text", "")
        changes = (message.text or "").strip()
        merged_text = base_text if changes.lower() in {"ок", "ok", "без изменений", "готово"} else _merge_raw_text(base_text, changes)
        company = companies[data["company_key"]]
        parsed = _prepare_parsed(merged_text)
        await _generate_and_send(
            message,
            storage,
            drafter,
            settings,
            data.get("doc_type", "door_offer"),
            company,
            parsed,
            merged_text,
            selected_counterparty_id=data.get("selected_counterparty_id"),
        )
        await state.clear()

    @router.callback_query(F.data.startswith("editdoc:"))
    async def edit_document(callback: CallbackQuery, state: FSMContext) -> None:
        _, document_id_text, operation = callback.data.split(":", 2)
        document = storage.get_document(callback.from_user.id, int(document_id_text))
        if not document:
            await _safe_callback_answer(callback, "Документ не найден", show_alert=True)
            return
        snapshot = document.get("snapshot") or {}
        if not snapshot:
            await _safe_callback_answer(callback, "Для этого документа нет данных для редактирования", show_alert=True)
            return
        await state.set_state(DocumentFlow.waiting_edit_value)
        await state.update_data(
            edit_document_id=document["id"],
            edit_operation=operation,
            edit_snapshot=snapshot,
            doc_type=document["doc_type"],
            company_key=document["company_key"],
            selected_counterparty_id=document.get("counterparty_id"),
        )
        await _show_callback_screen(callback, _edit_prompt(operation), reply_markup=_cancel_keyboard())
        await _safe_callback_answer(callback)

    @router.message(DocumentFlow.waiting_edit_value)
    async def receive_edit_value(message: Message, state: FSMContext) -> None:
        data = await state.get_data()
        snapshot = data.get("edit_snapshot") or {}
        fields = dict(snapshot.get("fields") or {})
        items = [dict(item) for item in (snapshot.get("items") or [])]
        operation = data.get("edit_operation")

        error = _apply_snapshot_edit(fields, items, operation, message.text or "", data.get("doc_type", ""))
        if error:
            await message.answer(error, reply_markup=_cancel_keyboard())
            return

        raw_text = _snapshot_to_raw_text({"fields": fields, "items": items})
        parsed = _prepare_parsed(raw_text)
        company = companies[data["company_key"]]
        await _generate_and_send(
            message,
            storage,
            drafter,
            settings,
            data["doc_type"],
            company,
            parsed,
            raw_text,
            selected_counterparty_id=data.get("selected_counterparty_id"),
        )
        await state.clear()

    @router.callback_query(F.data == "cancel")
    async def cancel(callback: CallbackQuery, state: FSMContext) -> None:
        await state.clear()
        await _show_callback_screen(callback, "Действие отменено.", reply_markup=_main_keyboard())
        await _safe_callback_answer(callback)

    @router.message(F.text)
    async def group_mention(message: Message, state: FSMContext) -> None:
        text = message.text or ""
        normalized = text.lower().strip()
        if message.chat.type in {"group", "supergroup"}:
            if normalized in {"id", "ид", "айди", "chat id", "чат id", "чат ид"}:
                await message.answer(_chat_id_text(message))
                return
            if normalized in {"menu", "меню", "док", "доки", "документы", "бот", "бот меню"}:
                await state.clear()
                await message.answer("Открыл меню документов для группы.", reply_markup=_main_keyboard())
                return
            if normalized.startswith(("товар ", "найди ", "поиск ")):
                query = normalized.split(maxsplit=1)[1]
                await message.answer(_product_search_text(query, moysklad), reply_markup=_catalog_keyboard())
                return
            if "@document07_bot" not in normalized and not normalized.startswith("бот"):
                return
            await state.clear()
            await message.answer("Я на связи. Открыл меню документов для группы.", reply_markup=_main_keyboard())
            return
        await message.answer("Нажмите /start или /menu, чтобы открыть меню.", reply_markup=_main_keyboard())

    dp = Dispatcher()
    dp.include_router(router)
    return dp


async def run_bot(settings: Settings) -> None:
    bot = Bot(token=settings.telegram_bot_token)
    dp = build_dispatcher(settings)
    await dp.start_polling(bot)


async def _prompt_quick_step(callback: CallbackQuery | None = None, message: Message | None = None, state: FSMContext | None = None) -> None:
    if state is None:
        return
    data = await state.get_data()
    normalized_index = _normalized_quick_index(data)
    if normalized_index != int(data.get("quick_step_index", 0)):
        await state.update_data(quick_step_index=normalized_index)
        data = await state.get_data()
    step = _current_quick_step(data)
    if not step:
        return

    prompt = step["prompt"]
    keyboard = _quick_step_keyboard(step, data)
    if callback is not None:
        await _show_callback_screen(callback, prompt, reply_markup=keyboard)
    elif message is not None:
        await message.answer(prompt, reply_markup=keyboard)


async def _consume_quick_value(
    state: FSMContext,
    step: dict[str, Any],
    raw_value: str,
    storage: Storage,
    drafter: OpenAIDrafter,
    settings: Settings,
    companies: dict[str, Company],
    callback: CallbackQuery | None = None,
    message: Message | None = None,
) -> None:
    data = await state.get_data()
    normalized_index = _normalized_quick_index(data)
    if normalized_index != int(data.get("quick_step_index", 0)):
        await state.update_data(quick_step_index=normalized_index)
        data = await state.get_data()
    fields = dict(data.get("draft_fields") or {})
    items = [dict(item) for item in (data.get("draft_items") or [])]
    value = raw_value.strip()

    if value == "skip":
        if step.get("required"):
            target = callback.message if callback else message
            if target is not None:
                await target.answer("Это поле обязательно. Заполните его или выберите другой режим.", reply_markup=_cancel_keyboard())
            return
    elif step["kind"] == "template":
        preset = TEMPLATE_PRESETS.get(value)
        if not preset:
            return
        fields["template"] = preset["template"]
        fields.setdefault("model", preset["model"])
    elif step["kind"] == "package":
        if value == "custom":
            items = []
        else:
            items = _package_items(step["doc_type"], value, fields)
    else:
        if value.lower() in {"-", "пропустить", "skip"} and not step.get("required"):
            value = "skip"
        if value == "skip":
            pass
        elif step["kind"] == "items":
            items = parse_items(value)
        else:
            field_name = step["field"]
            if field_name == "counterparty_address" and step["doc_type"] in DOOR_DOC_TYPES:
                fields["object_address"] = value
            fields[field_name] = value

    await state.update_data(draft_fields=fields, draft_items=items, quick_step_index=int(data.get("quick_step_index", 0)) + 1)

    next_step = _current_quick_step(await state.get_data())
    if next_step is None:
        target_message = message or (callback.message if callback else None)
        if target_message is not None:
            await _finalize_quick_flow(target_message, state, storage, drafter, settings, companies)
        return
    await _prompt_quick_step(callback=callback, message=message, state=state)


async def _finalize_quick_flow(
    message: Message,
    state: FSMContext,
    storage: Storage,
    drafter: OpenAIDrafter,
    settings: Settings,
    companies: dict[str, Company],
) -> None:
    data = await state.get_data()
    fields = dict(data.get("draft_fields") or {})
    items = [dict(item) for item in (data.get("draft_items") or [])]
    doc_type = data.get("doc_type", "")

    _apply_quick_defaults(doc_type, fields, items)
    raw_text = _snapshot_to_raw_text({"fields": fields, "items": items})
    parsed = _prepare_parsed(raw_text)

    if not _has_required_data(doc_type, parsed.fields, parsed.items):
        await message.answer(
            "Для этого документа все еще не хватает обязательных данных. Переключитесь на полный ввод или допишите недостающие поля.",
            reply_markup=_main_keyboard(),
        )
        await state.clear()
        return

    await state.clear()
    await message.answer(_quick_summary_text(doc_type, parsed.fields, parsed.items), reply_markup=_main_keyboard())

    # The generator uses the current state only for selected company and counterparty, so keep them before clear.
    company_key = data.get("company_key", "")
    selected_counterparty_id = data.get("selected_counterparty_id")
    if not company_key:
        await message.answer("Не выбрана организация. Нажмите /menu и попробуйте снова.", reply_markup=_main_keyboard())
        return
    company = companies[company_key]
    await _generate_and_send(
        message,
        storage,
        drafter,
        settings,
        doc_type,
        company,
        parsed,
        raw_text,
        selected_counterparty_id=selected_counterparty_id,
    )


async def _generate_and_send(
    message: Message,
    storage: Storage,
    drafter: OpenAIDrafter,
    settings: Settings,
    doc_type: str,
    company: Company,
    parsed,
    raw_text: str,
    selected_counterparty_id: int | None = None,
) -> None:
    ai_text = drafter.draft(doc_type, parsed)
    docx = generate_document(doc_type, company, parsed, settings.documents_dir, ai_text)
    snapshot = {"fields": parsed.fields, "items": parsed.items}
    counterparty_id = selected_counterparty_id or storage.upsert_counterparty(message.from_user.id, parsed.fields)

    try:
        pdf = generate_pdf_document(doc_type, company, parsed, settings.documents_dir, ai_text)
        document_id = storage.save_document(
            message.from_user.id,
            doc_type,
            company.key,
            pdf.filename,
            raw_text,
            snapshot=snapshot,
            counterparty_id=counterparty_id,
        )
        await message.answer_document(
            FSInputFile(pdf.path),
            caption=f"{pdf.title} готов. Номер в истории: #{document_id}. DOCX сохранен: {docx.filename}",
            reply_markup=_post_generate_keyboard(document_id),
        )
    except Exception as exc:
        document_id = storage.save_document(
            message.from_user.id,
            doc_type,
            company.key,
            docx.filename,
            raw_text,
            snapshot=snapshot,
            counterparty_id=counterparty_id,
        )
        await message.answer_document(
            FSInputFile(docx.path),
            caption=f"{docx.title} готов в DOCX. PDF не собрался: {exc}. Номер в истории: #{document_id}.",
            reply_markup=_post_generate_keyboard(document_id),
        )


async def _show_callback_screen(
    callback: CallbackQuery,
    text: str,
    reply_markup: InlineKeyboardMarkup,
) -> None:
    message = callback.message
    if message is None:
        return
    if message.text:
        try:
            await message.edit_text(text, reply_markup=reply_markup)
            return
        except TelegramBadRequest as exc:
            if "message is not modified" in str(exc).lower():
                await message.edit_reply_markup(reply_markup=reply_markup)
                return
    await message.answer(text, reply_markup=reply_markup)


def _main_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="🚪 КП по дверям", callback_data="doc:door_offer"), InlineKeyboardButton(text="💳 Счет с QR", callback_data="doc:invoice")],
            [InlineKeyboardButton(text="📐 Смета замера", callback_data="doc:measurement_estimate"), InlineKeyboardButton(text="📏 Акт замера", callback_data="doc:measurement_act")],
            [InlineKeyboardButton(text="📝 Договор поставки и монтажа", callback_data="doc:door_contract")],
            [InlineKeyboardButton(text="🔧 Акт монтажа", callback_data="doc:installation_act"), InlineKeyboardButton(text="✅ Акт работ", callback_data="doc:act")],
            [InlineKeyboardButton(text="🛡️ Гарантийный талон", callback_data="doc:warranty"), InlineKeyboardButton(text="⚖️ Ответ на претензию", callback_data="doc:claim_reply")],
            [InlineKeyboardButton(text="👤 Контрагенты", callback_data="counterparties"), InlineKeyboardButton(text="🔁 Повторить КП", callback_data="repeat_offer")],
            [InlineKeyboardButton(text="🗂 Каталог и цены", callback_data="catalog"), InlineKeyboardButton(text="⬇️ Скачать документы", callback_data="download_docs")],
            [InlineKeyboardButton(text="🕘 История", callback_data="history"), InlineKeyboardButton(text="📚 Все документы", callback_data="create")],
            [InlineKeyboardButton(text="🏦 Реквизиты ИП", callback_data="req:ip"), InlineKeyboardButton(text="🏢 Реквизиты ООО", callback_data="req:ooo")],
        ]
    )


def _doc_types_keyboard() -> InlineKeyboardMarkup:
    rows = []
    for key, title in DOC_TYPES.items():
        rows.append([InlineKeyboardButton(text=f"{DOC_ICONS.get(key, '📄')} {title}", callback_data=f"doc:{key}")])
    rows.append([InlineKeyboardButton(text="✖️ Отмена", callback_data="cancel")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def _catalog_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="🔎 Найти товар", callback_data="search_product"), InlineKeyboardButton(text="💰 Изменить цену", callback_data="edit_price")],
            [InlineKeyboardButton(text="🚪 КП по дверям", callback_data="doc:door_offer"), InlineKeyboardButton(text="💳 Счет с QR", callback_data="doc:invoice")],
            [InlineKeyboardButton(text="⬅️ Назад", callback_data="cancel")],
        ]
    )


def _companies_keyboard(companies: dict[str, Company], prefix: str) -> InlineKeyboardMarkup:
    rows = [[InlineKeyboardButton(text=f"{'🏦' if key == 'ip' else '🏢'} {company.label}", callback_data=f"{prefix}:{key}")] for key, company in companies.items()]
    rows.append([InlineKeyboardButton(text="✖️ Отмена", callback_data="cancel")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def _counterparties_keyboard(rows: list[dict[str, Any]], prefix: str) -> InlineKeyboardMarkup:
    keyboard = []
    for row in rows:
        suffix = f" / ИНН {row['inn']}" if row.get("inn") else ""
        keyboard.append([InlineKeyboardButton(text=f"👤 {row['name']}{suffix}", callback_data=f"{prefix}:{row['id']}")])
    keyboard.append([InlineKeyboardButton(text="⬅️ Назад", callback_data="cancel")])
    return InlineKeyboardMarkup(inline_keyboard=keyboard)


def _counterparties_menu_keyboard(rows: list[dict[str, Any]]) -> InlineKeyboardMarkup:
    keyboard: list[list[InlineKeyboardButton]] = [
        [InlineKeyboardButton(text="➕ Добавить", callback_data="cpadd"), InlineKeyboardButton(text="📥 Импорт", callback_data="cpimport")],
    ]
    for row in rows[:8]:
        suffix = f" / ИНН {row['inn']}" if row.get("inn") else ""
        keyboard.append([InlineKeyboardButton(text=f"👤 {row['name']}{suffix}", callback_data=f"cpview:{row['id']}")])
    keyboard.append([InlineKeyboardButton(text="⬅️ Назад", callback_data="cancel")])
    return InlineKeyboardMarkup(inline_keyboard=keyboard)


def _repeat_keyboard(rows: list[tuple[int, str, str, str]]) -> InlineKeyboardMarkup:
    keyboard = []
    for doc_id, doc_type, filename, created_at in rows:
        title = DOC_TYPES.get(doc_type, doc_type)
        keyboard.append([InlineKeyboardButton(text=f"🔁 #{doc_id} {title} от {created_at[:10]}", callback_data=f"rep:{doc_id}")])
    keyboard.append([InlineKeyboardButton(text="✖️ Отмена", callback_data="cancel")])
    return InlineKeyboardMarkup(inline_keyboard=keyboard)


def _download_keyboard(rows: list[tuple[int, str, str, str]]) -> InlineKeyboardMarkup:
    keyboard = []
    for doc_id, doc_type, filename, created_at in rows:
        title = DOC_TYPES.get(doc_type, doc_type)
        keyboard.append([InlineKeyboardButton(text=f"⬇️ #{doc_id} {title} от {created_at[:10]}", callback_data=f"dl:{doc_id}")])
    keyboard.append([InlineKeyboardButton(text="⬅️ Назад", callback_data="cancel")])
    return InlineKeyboardMarkup(inline_keyboard=keyboard)


def _cancel_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text="✖️ Отмена", callback_data="cancel")]])


def _mode_keyboard(doc_type: str, has_last: bool, has_counterparties: bool) -> InlineKeyboardMarkup:
    rows = []
    if doc_type == "invoice" and has_counterparties:
        rows.append([InlineKeyboardButton(text="👤 Выбрать контрагента из списка", callback_data="mode:counterparty")])
        rows.append([InlineKeyboardButton(text="⚡ Быстрый счет", callback_data="mode:quick"), InlineKeyboardButton(text="📝 Полный ввод", callback_data="mode:full")])
    else:
        rows.append([InlineKeyboardButton(text="⚡ Быстрый мастер", callback_data="mode:quick"), InlineKeyboardButton(text="📝 Полный ввод", callback_data="mode:full")])
    if has_last:
        rows.append([InlineKeyboardButton(text="♻️ Из прошлого документа", callback_data="mode:reuse")])
    if has_counterparties and doc_type != "invoice":
        rows.append([InlineKeyboardButton(text="👤 Из сохраненного контрагента", callback_data="mode:counterparty")])
    rows.append([InlineKeyboardButton(text="✖️ Отмена", callback_data="cancel")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def _post_generate_keyboard(document_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="👤 Поменять клиента", callback_data=f"editdoc:{document_id}:client"), InlineKeyboardButton(text="📍 Поменять адрес", callback_data=f"editdoc:{document_id}:address")],
            [InlineKeyboardButton(text="💰 Поменять цену", callback_data=f"editdoc:{document_id}:price"), InlineKeyboardButton(text="➕ Добавить позицию", callback_data=f"editdoc:{document_id}:add")],
            [InlineKeyboardButton(text="⬇️ Скачать документы", callback_data="download_docs"), InlineKeyboardButton(text="🏠 В меню", callback_data="cancel")],
        ]
    )


def _quick_step_keyboard(step: dict[str, Any], data: dict[str, Any]) -> InlineKeyboardMarkup:
    if step["kind"] == "template":
        rows = [
            [
                InlineKeyboardButton(text="Эконом", callback_data="quickv:economy"),
                InlineKeyboardButton(text="Стандарт", callback_data="quickv:standard"),
            ],
            [
                InlineKeyboardButton(text="Премиум", callback_data="quickv:premium"),
                InlineKeyboardButton(text="Межкомнатная", callback_data="quickv:interior"),
            ],
        ]
    elif step["kind"] == "package":
        if step["doc_type"] in {"installation_act", "act"}:
            rows = [
                [InlineKeyboardButton(text="Монтаж", callback_data="quickv:install"), InlineKeyboardButton(text="Монтаж + демонтаж", callback_data="quickv:install_demo")],
                [InlineKeyboardButton(text="Полный комплект", callback_data="quickv:full"), InlineKeyboardButton(text="Свои позиции", callback_data="quickv:custom")],
            ]
        else:
            rows = [
                [InlineKeyboardButton(text="Только дверь", callback_data="quickv:door"), InlineKeyboardButton(text="Дверь + доставка", callback_data="quickv:door_delivery")],
                [InlineKeyboardButton(text="Полный комплект", callback_data="quickv:full"), InlineKeyboardButton(text="Свои позиции", callback_data="quickv:custom")],
            ]
    else:
        rows = []

    if not step.get("required"):
        rows.append([InlineKeyboardButton(text="Пропустить", callback_data="quickv:skip")])
    rows.append([InlineKeyboardButton(text="✖️ Отмена", callback_data="cancel")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def _mode_prompt(doc_type: str, has_last: bool, has_counterparties: bool) -> str:
    if doc_type == "invoice":
        lines = [
            f"Документ: {DOC_TYPES.get(doc_type, doc_type)}",
            "",
            "Как оформить счет?",
        ]
        if has_counterparties:
            lines.append("1. Выбрать контрагента из списка и подставить его данные в счет.")
            lines.append("2. Быстрый счет: короткие шаги.")
            lines.append("3. Полный ввод: одно сообщение со всеми полями.")
        else:
            lines.append("1. Быстрый счет: короткие шаги.")
            lines.append("2. Полный ввод: одно сообщение со всеми полями.")
        if has_last:
            lines.append("4. Из прошлого документа: подтяну последние данные, вы пришлете только изменения.")
        return "\n".join(lines)

    lines = [
        f"Документ: {DOC_TYPES.get(doc_type, doc_type)}",
        "",
        "Как удобнее заполнить?",
        "1. Быстрый мастер: 3-6 коротких шагов.",
        "2. Полный ввод: одно сообщение со всеми полями.",
    ]
    if has_last:
        lines.append("3. Из прошлого документа: подтяну последние данные, вы пришлете только изменения.")
    if has_counterparties:
        lines.append("4. Из сохраненного контрагента: подставлю клиента, ИНН, телефон и адрес.")
    return "\n".join(lines)


def _details_prompt(doc_type: str = "", seeded_counterparty: bool = False) -> str:
    if doc_type == "door_offer":
        return (
            "Полный ввод для КП. Достаточно минимального набора:\n\n"
            "Клиент: Иванов Иван\n"
            "Адрес объекта: г. Уфа, ул. Ленина, 5\n"
            "Шаблон: стандарт\n"
            "Размер проема: 960 x 2050\n"
            "Позиции: стандарт | 1; доставка | 1\n\n"
            "Стандартный монтаж уже включен в стоимость двери. Опционально: телефон, цвет, открывание, скидка, срок, оплата."
        )
    if doc_type == "measurement_estimate":
        return (
            "Полный ввод для сметы. Минимум:\n\n"
            "Клиент: Иванов Иван\n"
            "Адрес объекта: г. Уфа, ул. Ленина, 5\n"
            "Размер проема: 960 x 2050\n"
            "Позиции: стандарт | 1\n\n"
            "Стандартный монтаж уже включен в стоимость двери. Опционально: шаблон, модель, толщина стены, цвет, скидка."
        )
    if doc_type in {"installation_act", "act"}:
        return (
            "Полный ввод для акта. Минимум:\n\n"
            "Клиент: Иванов Иван\n"
            "Адрес объекта: г. Уфа, ул. Ленина, 5\n"
            "Позиции: Монтаж входной двери | 1 | шт. | 6500\n\n"
            "Опционально: основание, срок, телефон, представитель."
        )
    if doc_type == "measurement_act":
        return (
            "Полный ввод для акта замера. Минимум:\n\n"
            "Клиент: Иванов Иван\n"
            "Адрес объекта: г. Уфа, ул. Ленина, 5\n"
            "Размер проема: 960 x 2050\n\n"
            "Опционально: модель, толщина стены, цвет, открывание, замерщик, комментарий."
        )
    if doc_type == "invoice":
        return (
            "Полный ввод для счета. Минимум:\n\n"
            "Клиент: ООО Ромашка\n"
            "Позиции: Входная дверь стандарт | 1 | шт. | 42000\n\n"
            "Опционально: ИНН, КПП, адрес, основание, НДС, назначение платежа."
        )
    return (
        "Пришлите данные одним сообщением. Теперь можно вводить только обязательный минимум, остальное опционально.\n\n"
        "Пример:\n"
        "Клиент: ООО Ромашка\n"
        "Предмет: Перевозка груза\n"
        "Позиции: Перевозка груза | 1 | рейс | 55000\n"
    )


def _help_text() -> str:
    return (
        "Команды:\n"
        "/start - главное меню\n"
        "/menu - меню в личке или группе\n"
        "/id - показать chat_id и user_id\n"
        "/requisites - реквизиты ИП/ООО\n"
        "/history - последние документы\n"
        "/counterparties - сохраненные контрагенты\n"
        "/catalog часть названия - поиск товара\n"
        "/product часть названия - поиск товара в каталоге и МойСклад\n"
        "/moysklad название или ИНН - поиск контрагента в МойСклад\n\n"
        "Для каждого документа есть быстрый мастер, полный ввод, подстановка из прошлого документа и сохраненные контрагенты."
    )


def _normalized_quick_index(data: dict[str, Any]) -> int:
    steps = _quick_steps(data.get("doc_type", ""))
    index = int(data.get("quick_step_index", 0))
    fields = dict(data.get("draft_fields") or {})
    items = list(data.get("draft_items") or [])
    while index < len(steps):
        step = steps[index]
        if step["kind"] == "items" and items and not step.get("required"):
            index += 1
            continue
        if step["kind"] in {"text", "template"} and fields.get(step["field"]):
            index += 1
            continue
        if step["kind"] == "package" and items:
            index += 1
            continue
        return index
    return index


def _current_quick_step(data: dict[str, Any]) -> dict[str, Any] | None:
    steps = _quick_steps(data.get("doc_type", ""))
    index = _normalized_quick_index(data)
    if index >= len(steps):
        return None
    return steps[index]


def _quick_steps(doc_type: str) -> list[dict[str, Any]]:
    if doc_type in DOOR_ORDER_DOC_TYPES:
        return [
            {"field": "counterparty_name", "prompt": "Кто клиент? Напишите имя или компанию.", "kind": "text", "required": True, "doc_type": doc_type},
            {"field": "counterparty_phone", "prompt": "Телефон клиента? Можно пропустить.", "kind": "text", "required": False, "doc_type": doc_type},
            {"field": "object_address", "prompt": "Адрес объекта?", "kind": "text", "required": True, "doc_type": doc_type},
            {"field": "template", "prompt": "Выберите шаблон двери:", "kind": "template", "required": True, "doc_type": doc_type},
            {"field": "opening_size", "prompt": "Размер проема? Например: 960 x 2050", "kind": "text", "required": True, "doc_type": doc_type},
            {"field": "items", "prompt": "Выберите комплект работ:", "kind": "package", "required": True, "doc_type": doc_type},
            {"field": "items", "prompt": "Если нужен свой состав работ, пришлите позиции. Иначе нажмите `Пропустить`.", "kind": "items", "required": False, "doc_type": doc_type},
            {"field": "payment_terms", "prompt": "Условия оплаты? Можно пропустить.", "kind": "text", "required": False, "doc_type": doc_type},
            {"field": "term", "prompt": "Срок выполнения? Можно пропустить.", "kind": "text", "required": False, "doc_type": doc_type},
        ]
    if doc_type == "measurement_act":
        return [
            {"field": "counterparty_name", "prompt": "Кто заказчик замера?", "kind": "text", "required": True, "doc_type": doc_type},
            {"field": "counterparty_phone", "prompt": "Телефон? Можно пропустить.", "kind": "text", "required": False, "doc_type": doc_type},
            {"field": "object_address", "prompt": "Адрес объекта?", "kind": "text", "required": True, "doc_type": doc_type},
            {"field": "template", "prompt": "Выберите тип двери:", "kind": "template", "required": True, "doc_type": doc_type},
            {"field": "opening_size", "prompt": "Размер проема?", "kind": "text", "required": True, "doc_type": doc_type},
            {"field": "wall_depth", "prompt": "Толщина стены? Можно пропустить.", "kind": "text", "required": False, "doc_type": doc_type},
            {"field": "description", "prompt": "Комментарий по замеру? Можно пропустить.", "kind": "text", "required": False, "doc_type": doc_type},
        ]
    if doc_type in {"installation_act", "act"}:
        return [
            {"field": "counterparty_name", "prompt": "Кто заказчик?", "kind": "text", "required": True, "doc_type": doc_type},
            {"field": "counterparty_phone", "prompt": "Телефон? Можно пропустить.", "kind": "text", "required": False, "doc_type": doc_type},
            {"field": "object_address" if doc_type == "installation_act" else "counterparty_address", "prompt": "Адрес объекта / адрес заказчика?", "kind": "text", "required": True, "doc_type": doc_type},
            {"field": "template", "prompt": "Выберите шаблон двери:", "kind": "template", "required": doc_type == "installation_act", "doc_type": doc_type},
            {"field": "items", "prompt": "Выберите состав работ:", "kind": "package", "required": True, "doc_type": doc_type},
            {"field": "items", "prompt": "Если нужен свой состав работ, пришлите позиции. Иначе нажмите `Пропустить`.", "kind": "items", "required": False, "doc_type": doc_type},
            {"field": "basis", "prompt": "Основание или номер заказа? Можно пропустить.", "kind": "text", "required": False, "doc_type": doc_type},
        ]
    if doc_type == "invoice":
        return [
            {"field": "counterparty_name", "prompt": "Кому выставляем счет?", "kind": "text", "required": True, "doc_type": doc_type},
            {"field": "counterparty_inn", "prompt": "ИНН? Можно пропустить.", "kind": "text", "required": False, "doc_type": doc_type},
            {"field": "counterparty_address", "prompt": "Адрес? Можно пропустить.", "kind": "text", "required": False, "doc_type": doc_type},
            {"field": "items", "prompt": "Напишите позиции в формате `Название | 1 | шт. | 42000`.", "kind": "items", "required": True, "doc_type": doc_type},
            {"field": "basis", "prompt": "Основание счета? Можно пропустить.", "kind": "text", "required": False, "doc_type": doc_type},
        ]
    return [
        {"field": "counterparty_name", "prompt": "Кто контрагент?", "kind": "text", "required": True, "doc_type": doc_type},
        {"field": "subject", "prompt": "Что нужно оформить? Напишите предмет документа.", "kind": "text", "required": True, "doc_type": doc_type},
        {"field": "items", "prompt": "Напишите позиции в формате `Название | 1 | усл. | 1000`.", "kind": "items", "required": True, "doc_type": doc_type},
    ]


def _package_items(doc_type: str, variant: str, fields: dict[str, str]) -> list[dict[str, str]]:
    model = fields.get("model") or TEMPLATE_PRESETS["standard"]["model"]
    is_interior = "межкомнат" in model.lower()
    install_name = "Монтаж межкомнатной двери" if is_interior else "Монтаж входной двери"

    if doc_type == "act":
        if variant == "install":
            return parse_items(f"Монтажные работы | 1 | усл. | 6500")
        if variant == "install_demo":
            return parse_items(f"Монтажные работы | 1 | усл. | 6500; Демонтаж старой двери | 1 | шт. | 1500")
        return parse_items(f"Монтажные работы | 1 | усл. | 6500; Дополнительные работы | 1 | усл. | 1500")

    if variant == "door":
        return parse_items(f"{model} | 1 | шт. | 0")
    if variant == "door_delivery":
        return parse_items(f"{model} | 1 | шт. | 0; Доставка по городу | 1 | усл. | 1500")
    if variant == "install":
        return parse_items(f"{install_name} | 1 | шт. | 6500")
    if variant == "install_demo":
        return parse_items(f"{install_name} | 1 | шт. | 6500; Демонтаж старой двери | 1 | шт. | 1500")
    if doc_type in DOOR_ORDER_DOC_TYPES:
        return parse_items(f"{model} | 1 | шт. | 0; Доставка по городу | 1 | усл. | 1500")
    return parse_items(f"{model} | 1 | шт. | 0; {install_name} | 1 | шт. | 6500; Доставка по городу | 1 | усл. | 1500")


def _apply_quick_defaults(doc_type: str, fields: dict[str, str], items: list[dict[str, str]]) -> None:
    if doc_type in DOOR_DOC_TYPES:
        fields.setdefault("opening_side", "правое наружное")
        fields.setdefault("color", "по согласованию")
    if doc_type in DOOR_ORDER_DOC_TYPES and not items:
        items.extend(_package_items(doc_type, "full", fields))
    if doc_type == "measurement_act":
        fields.setdefault("measurer", "Специалист исполнителя")
    if doc_type == "invoice":
        fields.setdefault("vat", "Без НДС")
    if doc_type in {"door_offer", "measurement_estimate"}:
        fields.setdefault("payment_terms", "Предоплата по счету")
        fields.setdefault("term", "по согласованию")


def _has_required_data(doc_type: str, fields: dict[str, str], items: list[dict[str, str]]) -> bool:
    if doc_type == "door_offer":
        return bool(fields.get("counterparty_name") and fields.get("object_address") and fields.get("opening_size") and items)
    if doc_type == "measurement_estimate":
        return bool(fields.get("counterparty_name") and fields.get("object_address") and fields.get("opening_size") and items)
    if doc_type == "measurement_act":
        return bool(fields.get("counterparty_name") and fields.get("object_address") and fields.get("opening_size"))
    if doc_type in {"installation_act", "act"}:
        return bool(fields.get("counterparty_name") and (fields.get("object_address") or fields.get("counterparty_address")) and items)
    if doc_type == "invoice":
        return bool(fields.get("counterparty_name") and items)
    return bool(fields.get("counterparty_name"))


def _prepare_parsed(raw_text: str):
    parsed = parse_user_input(raw_text)
    enrich_items_from_catalog(parsed)
    apply_discount(parsed)
    return parsed


def _merge_raw_text(base: str, extra: str) -> str:
    if not base.strip():
        return extra.strip()
    if not extra.strip():
        return base.strip()
    return f"{base.strip()}\n{extra.strip()}".strip()


def _snapshot_to_raw_text(snapshot: dict[str, Any]) -> str:
    fields = dict(snapshot.get("fields") or {})
    items = [dict(item) for item in (snapshot.get("items") or [])]
    lines: list[str] = []
    for key in FIELD_ORDER:
        value = str(fields.get(key) or "").strip()
        if value:
            lines.append(f"{FIELD_LABELS.get(key, key)}: {value}")
    remaining = [key for key in fields.keys() if key not in FIELD_ORDER]
    for key in remaining:
        value = str(fields.get(key) or "").strip()
        if value:
            lines.append(f"{FIELD_LABELS.get(key, key)}: {value}")
    visible_items = []
    for item in items:
        name = str(item.get("name", "")).strip()
        if not name or name.startswith("Скидка"):
            continue
        qty = str(item.get("qty", "1")).strip() or "1"
        unit = str(item.get("unit", "усл.")).strip() or "усл."
        price = str(item.get("price", item.get("total", "0"))).strip() or "0"
        visible_items.append(f"{name} | {qty} | {unit} | {price}")
    if visible_items:
        lines.append(f"Позиции: {'; '.join(visible_items)}")
    return "\n".join(lines)


def _counterparty_to_fields(row: dict[str, Any]) -> dict[str, str]:
    notes = dict(row.get("notes") or {})
    fields = {
        "counterparty_name": row.get("name", ""),
        "counterparty_inn": row.get("inn", ""),
        "counterparty_kpp": row.get("kpp", ""),
        "counterparty_ogrn": row.get("ogrn", ""),
        "counterparty_address": row.get("address", ""),
        "counterparty_phone": row.get("phone", ""),
        "counterparty_email": row.get("email", ""),
        "counterparty_manager": row.get("manager", ""),
    }
    if notes.get("object_address"):
        fields["object_address"] = notes["object_address"]
    return {key: value for key, value in fields.items() if value}


def _counterparty_text(row: dict[str, Any]) -> str:
    lines = [
        row["name"],
        f"ИНН: {row.get('inn') or '-'}",
        f"КПП: {row.get('kpp') or '-'}",
        f"ОГРН: {row.get('ogrn') or '-'}",
        f"Адрес: {row.get('address') or '-'}",
        f"Телефон: {row.get('phone') or '-'}",
        f"Email: {row.get('email') or '-'}",
        f"Представитель: {row.get('manager') or '-'}",
    ]
    return "\n".join(lines)


async def _download_text_file(bot: Bot, document: TgDocument) -> str | None:
    binary = await _download_binary_file(bot, document)
    if binary is None:
        return None
    try:
        return binary.decode("utf-8")
    except UnicodeDecodeError:
        try:
            return binary.decode("cp1251")
        except UnicodeDecodeError:
            return None


async def _download_binary_file(bot: Bot, document: TgDocument) -> bytes | None:
    buffer = io.BytesIO()
    try:
        await bot.download(document, destination=buffer)
    except Exception:
        return None
    return buffer.getvalue()


def _import_counterparties_from_text(storage: Storage, user_id: int, text: str, prefer_csv: bool = False) -> int:
    payloads = _parse_counterparty_import(text, prefer_csv=prefer_csv)
    saved = 0
    for payload in payloads:
        if storage.upsert_counterparty(user_id, payload):
            saved += 1
    return saved


def _import_counterparties_from_xlsx(storage: Storage, user_id: int, binary: bytes) -> int:
    payloads = _parse_counterparty_xlsx(binary)
    saved = 0
    for payload in payloads:
        if storage.upsert_counterparty(user_id, payload):
            saved += 1
    return saved


def _parse_counterparty_import(text: str, prefer_csv: bool = False) -> list[dict[str, str]]:
    stripped = text.strip()
    if not stripped:
        return []
    if prefer_csv or _looks_like_csv(stripped):
        rows = _parse_counterparty_csv(stripped)
        if rows:
            return rows
    blocks = [block.strip() for block in stripped.split("\n\n") if block.strip()]
    payloads: list[dict[str, str]] = []
    for block in blocks:
        parsed = parse_user_input(block)
        if parsed.fields.get("counterparty_name"):
            payloads.append(parsed.fields)
    if payloads:
        return payloads
    return _parse_counterparty_simple_lines(stripped)


def _looks_like_csv(text: str) -> bool:
    head = text.splitlines()[0].lower()
    return any(token in head for token in ["name", "inn", "kpp", "address", ";", ","])


def _parse_counterparty_csv(text: str) -> list[dict[str, str]]:
    sample = text.splitlines()[0] if text.splitlines() else ""
    delimiter = ";" if sample.count(";") > sample.count(",") else ","
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    payloads: list[dict[str, str]] = []
    for row in reader:
        payload = _map_counterparty_row(row)
        if payload.get("counterparty_name"):
            payloads.append(payload)
    return payloads


def _parse_counterparty_xlsx(binary: bytes) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except Exception:
        return []

    workbook = load_workbook(io.BytesIO(binary), read_only=True, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_normalize_counterparty_header(cell) for cell in rows[0]]
    payloads: list[dict[str, str]] = []
    for row in rows[1:]:
        mapping: dict[str, str] = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            value = row[index] if index < len(row) else None
            if value is None:
                continue
            mapping[header] = str(value).strip()
        payload = _map_counterparty_row(mapping)
        if payload.get("counterparty_name"):
            payloads.append(payload)
    return payloads


def _map_counterparty_row(row: dict[str, str]) -> dict[str, str]:
    payload = {
        "counterparty_name": _row_value(row, "name", "client", "counterparty_name", "контрагент", "клиент", "название"),
        "counterparty_inn": _row_value(row, "inn", "инн"),
        "counterparty_kpp": _row_value(row, "kpp", "кпп"),
        "counterparty_ogrn": _row_value(row, "ogrn", "огрн"),
        "counterparty_address": _row_value(row, "address", "counterparty_address", "адрес"),
        "counterparty_phone": _row_value(row, "phone", "telephone", "телефон"),
        "counterparty_email": _row_value(row, "email", "почта"),
        "counterparty_manager": _row_value(row, "manager", "representative", "представитель", "директор"),
    }
    return {key: value for key, value in payload.items() if value}


def _row_value(row: dict[str, str], *names: str) -> str:
    for name in names:
        value = row.get(name)
        if value:
            return value.strip()
    return ""


def _normalize_counterparty_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    return text.replace("ё", "е")


def _parse_counterparty_simple_lines(text: str) -> list[dict[str, str]]:
    payloads: list[dict[str, str]] = []
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        parts = [part.strip() for part in line.split(";")]
        if not parts:
            continue
        payload = {
            "counterparty_name": parts[0],
            "counterparty_inn": parts[1] if len(parts) > 1 else "",
            "counterparty_kpp": parts[2] if len(parts) > 2 else "",
            "counterparty_address": parts[3] if len(parts) > 3 else "",
            "counterparty_phone": parts[4] if len(parts) > 4 else "",
            "counterparty_email": parts[5] if len(parts) > 5 else "",
            "counterparty_manager": parts[6] if len(parts) > 6 else "",
        }
        payload = {key: value for key, value in payload.items() if value}
        if payload.get("counterparty_name"):
            payloads.append(payload)
    return payloads


def _quick_summary_text(doc_type: str, fields: dict[str, str], items: list[dict[str, str]]) -> str:
    lines = [f"Собрал данные для документа: {DOC_TYPES.get(doc_type, doc_type)}"]
    if fields.get("counterparty_name"):
        lines.append(f"Клиент: {fields['counterparty_name']}")
    if fields.get("object_address") or fields.get("counterparty_address"):
        lines.append(f"Адрес: {fields.get('object_address') or fields.get('counterparty_address')}")
    if items:
        lines.append("Позиции:")
        for item in items[:3]:
            lines.append(f"- {item.get('name', '')}")
    lines.append("Генерирую документ.")
    return "\n".join(lines)


def _edit_prompt(operation: str) -> str:
    if operation == "client":
        return "Напишите нового клиента. Можно просто имя или название компании."
    if operation == "address":
        return "Напишите новый адрес объекта / контрагента."
    if operation == "price":
        return "Напишите `название: цена` или просто число, если менять первую основную позицию."
    return "Напишите новую позицию в формате `Название | 1 | шт. | 6500`."


def _apply_snapshot_edit(fields: dict[str, str], items: list[dict[str, str]], operation: str, raw_value: str, doc_type: str) -> str | None:
    value = raw_value.strip()
    if not value:
        return "Пустое значение. Отправьте новое значение одним сообщением."

    if operation == "client":
        fields["counterparty_name"] = value
        return None

    if operation == "address":
        if doc_type in DOOR_DOC_TYPES or doc_type == "installation_act":
            fields["object_address"] = value
        fields["counterparty_address"] = value
        return None

    if operation == "add":
        items.extend(parse_items(value))
        return None

    target_items = [item for item in items if not str(item.get("name", "")).startswith("Скидка")]
    if not target_items:
        return "В документе нет позиций для изменения цены."

    target = target_items[0]
    if ":" in value:
        name_query, raw_price = value.split(":", 1)
        raw_price = raw_price.strip()
        for item in target_items:
            if name_query.strip().lower() in str(item.get("name", "")).lower():
                target = item
                break
    else:
        raw_price = value

    price = _decimal(raw_price)
    if price <= 0:
        return "Цена должна быть больше нуля."
    qty = _decimal(str(target.get("qty", "1")))
    target["price"] = f"{price.quantize(Decimal('0.01'))}"
    target["total"] = f"{(qty * price).quantize(Decimal('0.01'))}"
    return None


def _parse_price(value: str) -> int:
    clean = "".join(ch for ch in value if ch.isdigit())
    return int(clean or "0")


def _product_search_text(query: str, moysklad: MoySkladClient) -> str:
    text = search_text(query)
    if not moysklad.enabled:
        return text
    try:
        rows = moysklad.search_assortment(query)
    except Exception as exc:
        return text + f"\n\nМойСклад: ошибка поиска товара: {exc}"
    if not rows:
        return text + "\n\nМойСклад: товар не найден."

    lines = [text, "", "МойСклад:"]
    for row in rows:
        price = _moysklad_sale_price(row)
        lines.append(f"- {row.get('name', '-')}: {price}")
    return "\n".join(lines)


def _moysklad_sale_price(row: dict) -> str:
    prices = row.get("salePrices") or []
    if not prices:
        return "цена не указана"
    value = prices[0].get("value")
    if value is None:
        return "цена не указана"
    return f"{value / 100:.2f} руб."


def _command_args(message: Message) -> str:
    text = message.text or ""
    parts = text.split(maxsplit=1)
    if len(parts) == 1:
        return ""
    return parts[1].strip()


def _chat_id_text(message: Message) -> str:
    return (
        "ID для этой переписки:\n"
        f"chat_id: {message.chat.id}\n"
        f"chat_type: {message.chat.type}\n"
        f"user_id: {message.from_user.id if message.from_user else '-'}"
    )


async def _safe_callback_answer(callback: CallbackQuery, text: str | None = None, show_alert: bool = False) -> None:
    try:
        if text is None:
            await callback.answer()
        else:
            await callback.answer(text, show_alert=show_alert)
    except Exception:
        return


def _decimal(value: str) -> Decimal:
    try:
        clean = "".join(ch for ch in str(value).replace(",", ".") if ch.isdigit() or ch in ".-")
        return Decimal(clean or "0")
    except Exception:
        return Decimal("0")
